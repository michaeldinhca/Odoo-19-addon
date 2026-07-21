import base64
import io

import openpyxl
from dateutil.relativedelta import relativedelta

from odoo import fields
from odoo.exceptions import UserError
from odoo.tests.common import TransactionCase, tagged

# See test_cash_flow_forecast_report.py for why dates are relative to
# today, why taxes_id is cleared on the test product, and why flush_all()
# is called before anything that reads cash.flow.forecast.report (directly,
# or indirectly via scenario.action_run()) right after posting/confirming -
# that raw SQL view has no ORM-visible dependency on account.move/sale.order
# to trigger an automatic flush.


@tagged('post_install', '-at_install')
class TestCashFlowForecastScenario(TransactionCase):

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.partner = cls.env['res.partner'].create({'name': 'Scenario Test Partner'})
        cls.product = cls.env['product.product'].create({
            'name': 'Scenario Test Product',
            'type': 'consu',
            'list_price': 100.0,
            'taxes_id': [(6, 0, [])],
            'supplier_taxes_id': [(6, 0, [])],
        })
        cls.today = fields.Date.context_today(cls.env['account.move'])

        cls.move_near = cls.env['account.move'].create({
            'move_type': 'out_invoice',
            'partner_id': cls.partner.id,
            'invoice_date': cls.today,
            'invoice_date_due': cls.today + relativedelta(days=15),
            'invoice_line_ids': [(0, 0, {
                'product_id': cls.product.id,
                'quantity': 1,
                'price_unit': 500.0,
            })],
        })
        cls.move_near.action_post()
        cls.env.flush_all()

        # Beyond the 30-day horizon used below, but within the 180-day one.
        cls.move_far = cls.env['account.move'].create({
            'move_type': 'out_invoice',
            'partner_id': cls.partner.id,
            'invoice_date': cls.today,
            'invoice_date_due': cls.today + relativedelta(days=90),
            'invoice_line_ids': [(0, 0, {
                'product_id': cls.product.id,
                'quantity': 1,
                'price_unit': 700.0,
            })],
        })
        cls.move_far.action_post()
        cls.env.flush_all()

    def test_run_builds_lines_matching_live_report_within_horizon(self):
        scenario = self.env['cash.flow.forecast.scenario'].create({
            'name': 'Test Scenario',
            'method': 'confirmed_documents',
            'horizon': '30',
        })
        scenario.action_run()

        self.assertTrue(scenario.computed_on)
        self.assertIn(self.move_near.id, scenario.line_ids.mapped('move_id.id'))
        self.assertNotIn(self.move_far.id, scenario.line_ids.mapped('move_id.id'))

        live_rows = self.env['cash.flow.forecast.report'].search([
            ('company_id', '=', scenario.company_id.id),
            ('bucket_date', '<=', self.today + relativedelta(days=30)),
        ])
        self.assertEqual(len(scenario.line_ids), len(live_rows))

    def test_rerun_with_wider_horizon_deletes_and_rebuilds(self):
        scenario = self.env['cash.flow.forecast.scenario'].create({
            'name': 'Test Scenario',
            'method': 'confirmed_documents',
            'horizon': '30',
        })
        scenario.action_run()
        first_run_line_ids = set(scenario.line_ids.ids)
        self.assertNotIn(self.move_far.id, scenario.line_ids.mapped('move_id.id'))

        scenario.horizon = '180'
        scenario.action_run()

        self.assertFalse(set(scenario.line_ids.ids) & first_run_line_ids)
        self.assertIn(self.move_far.id, scenario.line_ids.mapped('move_id.id'))

    def test_view_url_set_only_for_lines_with_a_source_record(self):
        scenario = self.env['cash.flow.forecast.scenario'].create({
            'name': 'Test Scenario',
            'method': 'confirmed_documents',
            'horizon': '90',
        })
        scenario.action_run()

        invoice_line = scenario.line_ids.filtered(lambda l: l.move_id == self.move_near)
        self.assertTrue(invoice_line.view_url)
        self.assertIn(str(self.move_near.id), invoice_line.view_url)

        opening_line = scenario.line_ids.filtered(lambda l: l.source_type == 'opening_balance')
        self.assertFalse(opening_line.view_url)

    def test_description_includes_installment_ordinal_and_overdue_marker(self):
        # invoice_date is in the past so the first (shorter) installment's
        # due date has already passed, without needing a negative nb_days.
        term = self.env['account.payment.term'].create({
            'name': 'Scenario Test 50-50 Term',
            'line_ids': [
                (0, 0, {'value': 'percent', 'value_amount': 50, 'nb_days': 5}),
                (0, 0, {'value': 'percent', 'value_amount': 50, 'nb_days': 40}),
            ],
        })
        split_move = self.env['account.move'].create({
            'move_type': 'out_invoice',
            'partner_id': self.partner.id,
            'invoice_date': self.today - relativedelta(days=20),
            'invoice_payment_term_id': term.id,
            'invoice_line_ids': [(0, 0, {
                'product_id': self.product.id,
                'quantity': 1,
                'price_unit': 1000.0,
            })],
        })
        split_move.action_post()
        self.env.flush_all()

        scenario = self.env['cash.flow.forecast.scenario'].create({
            'name': 'Test Scenario',
            'method': 'confirmed_documents',
            'horizon': '90',
        })
        scenario.action_run()

        split_lines = scenario.line_ids.filtered(lambda l: l.move_id == split_move).sorted('sequence')
        self.assertEqual(len(split_lines), 2)
        self.assertIn('installment #1', split_lines[0].name)
        self.assertIn('installment #2', split_lines[1].name)
        # The first installment's due date (today - 15 days) is already overdue.
        self.assertTrue(split_lines[0].is_overdue)
        self.assertIn('Overdue', split_lines[0].name)
        self.assertFalse(split_lines[1].is_overdue)
        self.assertNotIn('Overdue', split_lines[1].name)

        single_line = scenario.line_ids.filtered(lambda l: l.move_id == self.move_near)
        self.assertNotIn('installment', single_line.name)

    def test_description_prefixes_invoice_origin_matching_odoo_native_style(self):
        # Odoo's own journal-item labels for a split-installment invoice
        # look like "S00026 - INV/2026/00012 installment #1" - match that.
        order = self.env['sale.order'].create({
            'partner_id': self.partner.id,
            'order_line': [(0, 0, {
                'product_id': self.product.id,
                'product_uom_qty': 1,
                'price_unit': 900.0,
            })],
        })
        order.action_confirm()
        invoices = order._create_invoices()
        invoices.invoice_date = self.today
        invoices.action_post()
        self.env.flush_all()

        scenario = self.env['cash.flow.forecast.scenario'].create({
            'name': 'Test Scenario',
            'method': 'confirmed_documents',
            'horizon': '30',
        })
        scenario.action_run()

        line = scenario.line_ids.filtered(lambda l: l.move_id == invoices)
        self.assertEqual(len(line), 1)
        self.assertTrue(line.name.startswith(order.name), line.name)
        self.assertIn(invoices.name, line.name)
        self.assertNotIn('installment', line.name)

    def test_opening_balance_override_shifts_running_balance(self):
        plain_scenario = self.env['cash.flow.forecast.scenario'].create({
            'name': 'Plain Scenario',
            'method': 'confirmed_documents',
            'horizon': '30',
        })
        plain_scenario.action_run()
        plain_opening = plain_scenario.line_ids.filtered(lambda l: l.source_type == 'opening_balance')
        plain_first_invoice_line = plain_scenario.line_ids.filtered(lambda l: l.move_id == self.move_near)

        override_scenario = self.env['cash.flow.forecast.scenario'].create({
            'name': 'Override Scenario',
            'method': 'confirmed_documents',
            'horizon': '30',
            'opening_balance_override': 10000.0,
        })
        override_scenario.action_run()
        override_opening = override_scenario.line_ids.filtered(lambda l: l.source_type == 'opening_balance')
        override_first_invoice_line = override_scenario.line_ids.filtered(lambda l: l.move_id == self.move_near)

        self.assertAlmostEqual(override_opening.net_amount, 10000.0, places=2)
        self.assertAlmostEqual(override_opening.running_balance, 10000.0, places=2)
        self.assertIn('Manual Override', override_opening.name)
        self.assertIn('Computed from Bank/Cash', plain_opening.name)

        # Every later row's running balance must shift by the same delta as
        # the opening balance itself, not just the opening balance row.
        delta = override_opening.net_amount - plain_opening.net_amount
        self.assertAlmostEqual(
            override_first_invoice_line.running_balance,
            plain_first_invoice_line.running_balance + delta,
            places=2,
        )

    def test_opening_balance_description_names_the_contributing_journal(self):
        bank_journal = self.env['account.journal'].search(
            [('type', '=', 'bank'), ('company_id', '=', self.env.company.id)], limit=1
        )
        income_account = self.env['account.account'].search(
            [('account_type', '=', 'income'), ('company_ids', 'in', self.env.company.id)], limit=1
        )
        move = self.env['account.move'].create({
            'move_type': 'entry',
            'journal_id': bank_journal.id,
            'line_ids': [
                (0, 0, {'account_id': bank_journal.default_account_id.id, 'debit': 3000.0, 'credit': 0.0, 'name': 'Test bank entry'}),
                (0, 0, {'account_id': income_account.id, 'debit': 0.0, 'credit': 3000.0, 'name': 'Test bank entry'}),
            ],
        })
        move.action_post()
        self.env.flush_all()

        scenario = self.env['cash.flow.forecast.scenario'].create({
            'name': 'Test Scenario',
            'method': 'confirmed_documents',
            'horizon': '30',
        })
        scenario.action_run()

        opening_line = scenario.line_ids.filtered(lambda l: l.source_type == 'opening_balance')
        self.assertIn(bank_journal.name, opening_line.name)

    def test_export_xlsx_includes_header_info_and_lines(self):
        scenario = self.env['cash.flow.forecast.scenario'].create({
            'name': 'Export Test Scenario',
            'method': 'confirmed_documents',
            'horizon': '30',
        })
        scenario.action_run()

        action = scenario.action_export_xlsx()
        self.assertEqual(action['type'], 'ir.actions.act_url')
        attachment_id = int(action['url'].rsplit('/', 1)[-1])
        attachment = self.env['ir.attachment'].browse(attachment_id)
        self.assertTrue(attachment.exists())

        content = base64.b64decode(attachment.datas)
        workbook = openpyxl.load_workbook(io.BytesIO(content))
        sheet = workbook.active

        self.assertEqual(sheet.cell(row=1, column=1).value, 'Scenario')
        self.assertEqual(sheet.cell(row=1, column=2).value, 'Export Test Scenario')
        self.assertEqual(sheet.cell(row=2, column=1).value, 'Method')
        self.assertEqual(sheet.cell(row=3, column=1).value, 'Forecast Horizon')
        self.assertEqual(sheet.cell(row=3, column=2).value, 'Next 30 Days')
        self.assertEqual(sheet.cell(row=4, column=1).value, 'Computed On')

        header_row = 6
        self.assertEqual(sheet.cell(row=header_row, column=1).value, 'Forecast Date')
        self.assertEqual(sheet.cell(row=header_row, column=8).value, 'Forecasted Balance')
        # At least one data row beyond the header
        self.assertIsNotNone(sheet.cell(row=header_row + 1, column=1).value)

    def _post_bank_entry(self, date, amount, bank_account=None, income_account=None, journal=None, company=None):
        """Posts a direct entry crediting/debiting the bank account by
        `amount` (positive = incoming, negative = outgoing) on `date` -
        same helper pattern as test_cash_flow_actual_daily.py. Pass explicit
        bank_account/income_account/journal/company to post under an
        isolated test company instead of the shared default one.
        """
        bank_account = bank_account or self.env['account.journal'].search(
            [('type', '=', 'bank'), ('company_id', '=', self.env.company.id)], limit=1
        ).default_account_id
        income_account = income_account or self.env['account.account'].search(
            [('account_type', '=', 'income'), ('company_ids', 'in', self.env.company.id)], limit=1
        )
        debit_account, credit_account = (
            (bank_account, income_account) if amount >= 0 else (income_account, bank_account)
        )
        move_vals = {
            'move_type': 'entry',
            'date': date,
            'line_ids': [
                (0, 0, {'account_id': debit_account.id, 'debit': abs(amount), 'credit': 0.0, 'name': 'Test'}),
                (0, 0, {'account_id': credit_account.id, 'debit': 0.0, 'credit': abs(amount), 'name': 'Test'}),
            ],
        }
        if journal is not None:
            move_vals['journal_id'] = journal.id
        if company is not None:
            move_vals['company_id'] = company.id
        move = self.env['account.move'].create(move_vals)
        move.action_post()
        return move

    def _create_isolated_bank_company(self):
        """Creates a throwaway company plus a bank/income account and bank
        journal for it. The historical-trend method aggregates ALL cash
        activity for a company within a date range (unlike every other test
        in this file, which filters to specific record ids) - so it isn't
        immune to leftover account.move.line rows from earlier manual
        shell-based debugging sessions run against this same dev database.
        A brand new company is guaranteed to have zero such history.
        """
        company = self.env['res.company'].create({'name': 'Trend Isolation Co'})
        self.env.user.write({'company_ids': [(4, company.id)]})
        bank_account = self.env['account.account'].create({
            'name': 'Trend Isolation Bank',
            'code': 'TIB1',
            'account_type': 'asset_cash',
            'company_ids': [(6, 0, [company.id])],
        })
        income_account = self.env['account.account'].create({
            'name': 'Trend Isolation Income',
            'code': 'TII1',
            'account_type': 'income',
            'company_ids': [(6, 0, [company.id])],
        })
        bank_journal = self.env['account.journal'].create({
            'name': 'Trend Isolation Bank Journal',
            'code': 'TIBJ',
            'type': 'bank',
            'company_id': company.id,
            'default_account_id': bank_account.id,
        })
        return company, bank_account, income_account, bank_journal

    def test_historical_trend_computes_weekly_average_and_range(self):
        company, bank_account, income_account, bank_journal = self._create_isolated_bank_company()

        # 4 distinct weeks (exactly 7 days apart, so each reliably falls in
        # a different ISO week regardless of what day "today" is), with
        # deliberately varying amounts so mean != min != max - proves the
        # range calculation isn't just returning a constant.
        incoming_by_week_back = {1: 1000.0, 2: 2000.0, 3: 1000.0, 4: 1500.0}
        outgoing_by_week_back = {1: 500.0, 2: 700.0, 3: 500.0, 4: 600.0}
        for weeks_back, amount in incoming_by_week_back.items():
            self._post_bank_entry(
                self.today - relativedelta(weeks=weeks_back), amount,
                bank_account=bank_account, income_account=income_account,
                journal=bank_journal, company=company,
            )
        for weeks_back, amount in outgoing_by_week_back.items():
            self._post_bank_entry(
                self.today - relativedelta(weeks=weeks_back), -amount,
                bank_account=bank_account, income_account=income_account,
                journal=bank_journal, company=company,
            )
        self.env.flush_all()

        scenario = self.env['cash.flow.forecast.scenario'].create({
            'name': 'Trend Test Scenario',
            'method': 'historical_trend',
            'horizon': '30',
            'trend_lookback_weeks': '4',
            'opening_balance_override': 5000.0,
            'company_id': company.id,
        })
        scenario.action_run()

        all_lines = scenario.trend_line_ids.sorted('sequence')
        historical = all_lines.filtered(lambda l: l.row_type == 'historical')
        opening = all_lines.filtered(lambda l: l.row_type == 'opening_balance')
        projected = all_lines.filtered(lambda l: l.row_type == 'projected')

        expected_weeks = 5  # ceil(30 / 7)
        self.assertEqual(len(projected), expected_weeks)
        self.assertEqual(len(historical), 4)
        self.assertEqual(len(opening), 1)

        # Row order: historical (oldest->newest), then opening, then projected.
        self.assertEqual(list(all_lines), list(historical) + list(opening) + list(projected))

        first = projected[0]
        self.assertEqual(first.week_start_date, self.today)
        self.assertEqual(first.week_end_date, self.today + relativedelta(days=6))
        self.assertAlmostEqual(first.amount_in_avg, 1375.0, places=2)  # (1000+2000+1000+1500)/4
        self.assertAlmostEqual(first.amount_in_min, 1000.0, places=2)
        self.assertAlmostEqual(first.amount_in_max, 2000.0, places=2)
        self.assertAlmostEqual(first.amount_out_avg, 575.0, places=2)  # (500+700+500+600)/4
        self.assertAlmostEqual(first.amount_out_min, 500.0, places=2)
        self.assertAlmostEqual(first.amount_out_max, 700.0, places=2)
        self.assertAlmostEqual(first.net_avg, 800.0, places=2)  # 1375 - 575
        self.assertAlmostEqual(first.running_balance_avg, 5800.0, places=2)  # 5000 + 800

        # Running balance accumulates the same net_avg each subsequent week.
        second = projected[1]
        self.assertAlmostEqual(second.running_balance_avg, 6600.0, places=2)  # 5800 + 800
        self.assertEqual(second.week_start_date, self.today + relativedelta(weeks=1))

        # Opening balance row: same shared helper/number as Confirmed Documents.
        self.assertAlmostEqual(opening.running_balance_avg, 5000.0, places=2)
        self.assertIn('Manual Override', opening.name)

        # Historical rows carry the week's ACTUAL totals (min==avg==max, no
        # range for a single realized week), and a real reconstructed
        # running balance walked backward from the opening balance - not
        # the averaged figures used for the projected rows.
        newest = historical[3]  # 1 week back: in=1000, out=500, net=500
        oldest = historical[0]  # 4 weeks back: in=1500, out=600, net=900
        self.assertAlmostEqual(newest.amount_in_avg, 1000.0, places=2)
        self.assertAlmostEqual(newest.amount_in_min, 1000.0, places=2)
        self.assertAlmostEqual(newest.amount_in_max, 1000.0, places=2)
        self.assertAlmostEqual(newest.net_avg, 500.0, places=2)
        # Newest historical week's ending balance == today's true balance,
        # since nothing happened between the end of that week and today.
        self.assertAlmostEqual(newest.running_balance_avg, 5000.0, places=2)
        # Walking backward: 5000 - 500 (week 1 net) = 4500 (week 2's end),
        # -1300 (week 2 net) = 3200 (week 3's end), -500 (week 3 net) = 2700.
        self.assertAlmostEqual(oldest.running_balance_avg, 2700.0, places=2)

    def test_biweekly_salary_added_to_confirmed_documents(self):
        # Compares a plain run against a salary-added run on the same
        # (default, possibly not pristine) company rather than asserting
        # absolute balances - the shared dev database can carry leftover
        # real documents from manual verification sessions, same lesson as
        # the historical-trend test isolation fix. A delta comparison at a
        # known point (move_near, always present) is immune to that.
        plain_scenario = self.env['cash.flow.forecast.scenario'].create({
            'name': 'No Salary Scenario',
            'method': 'confirmed_documents',
            'horizon': '30',
        })
        plain_scenario.action_run()
        plain_move_line = plain_scenario.line_ids.filtered(lambda l: l.move_id == self.move_near)

        salary_scenario = self.env['cash.flow.forecast.scenario'].create({
            'name': 'Salary Test Scenario',
            'method': 'confirmed_documents',
            'horizon': '30',
            'biweekly_salary_amount': 400.0,
            'biweekly_salary_start_date': self.today + relativedelta(days=5),
        })
        salary_scenario.action_run()

        # Occurrences every 14 days from today+5 through today+30: today+5
        # and today+19 (today+33 falls outside the 30-day horizon).
        salary_lines = salary_scenario.line_ids.filtered(lambda l: l.source_type == 'salary').sorted('sequence')
        self.assertEqual(len(salary_lines), 2)
        self.assertEqual(salary_lines[0].bucket_date, self.today + relativedelta(days=5))
        self.assertEqual(salary_lines[1].bucket_date, self.today + relativedelta(days=19))
        for line in salary_lines:
            self.assertAlmostEqual(line.amount_out, 400.0, places=2)
            self.assertEqual(line.name, 'Bi-Weekly Salary')

        # move_near (due today+15) sits between the two salary occurrences,
        # so exactly one salary payment (today+5) precedes it - its running
        # balance must be exactly one salary payment lower than without the
        # salary setting, regardless of whatever else is in the forecast.
        salary_move_line = salary_scenario.line_ids.filtered(lambda l: l.move_id == self.move_near)
        self.assertAlmostEqual(
            salary_move_line.running_balance, plain_move_line.running_balance - 400.0, places=2,
        )

    def test_biweekly_salary_start_date_in_past_recurs_forward_from_next_occurrence(self):
        scenario = self.env['cash.flow.forecast.scenario'].create({
            'name': 'Salary Past Anchor Scenario',
            'method': 'confirmed_documents',
            'horizon': '30',
            'biweekly_salary_amount': 300.0,
            'biweekly_salary_start_date': self.today - relativedelta(days=5),
        })
        scenario.action_run()

        # Anchor is 5 days in the past - the first occurrence on/after today
        # is 9 days forward (14 - 5), then every 14 days after that: today+9
        # and today+23 (today+37 falls outside the 30-day horizon).
        salary_lines = scenario.line_ids.filtered(lambda l: l.source_type == 'salary').sorted('sequence')
        self.assertEqual(len(salary_lines), 2)
        self.assertEqual(salary_lines[0].bucket_date, self.today + relativedelta(days=9))
        self.assertEqual(salary_lines[1].bucket_date, self.today + relativedelta(days=23))

    def test_biweekly_salary_folds_into_historical_trend_projected_weeks(self):
        company, bank_account, income_account, bank_journal = self._create_isolated_bank_company()

        for weeks_back, amount in {1: 1000.0, 2: 1000.0}.items():
            self._post_bank_entry(
                self.today - relativedelta(weeks=weeks_back), amount,
                bank_account=bank_account, income_account=income_account,
                journal=bank_journal, company=company,
            )
        for weeks_back, amount in {1: 200.0, 2: 300.0}.items():
            self._post_bank_entry(
                self.today - relativedelta(weeks=weeks_back), -amount,
                bank_account=bank_account, income_account=income_account,
                journal=bank_journal, company=company,
            )
        self.env.flush_all()

        scenario = self.env['cash.flow.forecast.scenario'].create({
            'name': 'Trend Salary Scenario',
            'method': 'historical_trend',
            'horizon': '30',
            'trend_lookback_weeks': '4',
            'opening_balance_override': 5000.0,
            'biweekly_salary_amount': 400.0,
            'biweekly_salary_start_date': self.today,
            'company_id': company.id,
        })
        scenario.action_run()

        # avg_in = 1000, avg_out = (200+300)/2 = 250, min_out=200, max_out=300.
        # Salary occurrences land on today (week 0), today+14 (week 2) and
        # today+28 (week 4) - weeks 1 and 3 are untouched.
        projected = scenario.trend_line_ids.filtered(lambda l: l.row_type == 'projected').sorted('sequence')
        self.assertEqual(len(projected), 5)

        week0 = projected[0]
        self.assertAlmostEqual(week0.amount_out_avg, 650.0, places=2)  # 250 + 400
        self.assertAlmostEqual(week0.amount_out_min, 600.0, places=2)  # 200 + 400
        self.assertAlmostEqual(week0.amount_out_max, 700.0, places=2)  # 300 + 400
        self.assertAlmostEqual(week0.net_avg, 350.0, places=2)  # 1000 - 650
        self.assertEqual(week0.name, 'Includes Bi-Weekly Salary')
        self.assertAlmostEqual(week0.running_balance_avg, 5350.0, places=2)  # 5000 + 350

        week1 = projected[1]
        self.assertAlmostEqual(week1.amount_out_avg, 250.0, places=2)  # no salary this week
        self.assertFalse(week1.name)
        self.assertAlmostEqual(week1.running_balance_avg, 6100.0, places=2)  # 5350 + (1000 - 250)

        week2 = projected[2]
        self.assertAlmostEqual(week2.amount_out_avg, 650.0, places=2)
        self.assertEqual(week2.name, 'Includes Bi-Weekly Salary')
        self.assertAlmostEqual(week2.running_balance_avg, 6450.0, places=2)  # 6100 + 350

    def test_historical_trend_requires_at_least_two_weeks_of_history(self):
        company, _bank_account, _income_account, _bank_journal = self._create_isolated_bank_company()
        scenario = self.env['cash.flow.forecast.scenario'].create({
            'name': 'Trend Test Scenario No History',
            'method': 'historical_trend',
            'horizon': '30',
            'trend_lookback_weeks': '4',
            'company_id': company.id,
        })
        with self.assertRaises(UserError):
            scenario.action_run()
